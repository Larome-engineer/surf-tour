import logging
from io import BytesIO
from openpyxl import Workbook
from tortoise import Tortoise

logger = logging.getLogger(__name__)


async def export_all_models_to_excel() -> BytesIO:
    wb = Workbook()
    del wb["Sheet"]

    models = Tortoise.apps.get("models")
    if not models:
        logger.error("Tortoise models not loaded")
        return BytesIO()

    for idx, (name, model) in enumerate(models.items()):
        if model is None:
            continue

        sheet_name = model.__name__ or f"Sheet_{idx}"
        sheet_name = str(sheet_name).replace("/", "_")[:31]
        original = sheet_name
        i = 1
        while sheet_name in wb.sheetnames:
            sheet_name = f"{original}_{i}"
            i += 1

        ws = wb.create_sheet(title=sheet_name)

        try:
            records = await model.all().prefetch_related()
            if not records:
                continue

            # Берём реальные поля из первого объекта
            first_obj = records[0]
            # Берём только настоящие поля (без служебных)
            fields = [
                f for f in first_obj.__dict__.keys()
                if not f.startswith("_")
            ]

            ws.append(fields)

            for obj in records:
                row = []
                for f in fields:
                    val = getattr(obj, f, None)
                    if callable(val):
                        row.append("")
                    else:
                        try:
                            row.append(str(val) if val is not None else "")
                        except:
                            row.append("")
                ws.append(row)

        except Exception as e:
            logger.error(f"❌ Failed model {model.__name__}: {e}")
            continue

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
